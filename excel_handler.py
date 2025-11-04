import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List


class ExcelHandler:
    """
    Класс для работы с Excel файлом настроек
    """
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = None
        self.settings_sheet = None
        self.work_sheet = None
        self.translator_sheet = None
        
    def load_workbook(self):
        """Загружает Excel файл"""
        try:
            self.wb = openpyxl.load_workbook(self.excel_path)
            self.settings_sheet = self.wb['Настройки']
            self.work_sheet = self.wb['Рабочий']
            self.translator_sheet = self.wb['Переводчик']
            return True
        except Exception as e:
            print(f"Ошибка при загрузке Excel файла: {e}")
            return False

    def get_settings(self) -> Dict:
        """
        Получает настройки с листа 'Настройки'
        
        Returns:
            Dict: Словарь с настройками
        """
        return {
            'exchange_rate': float(self.settings_sheet['B1'].value),  # Курс евро
            'margin_percent': float(self.settings_sheet['B2'].value),  # Наценка %
            'item_prefix': self.settings_sheet['B3'].value,           # Префикс
            'start_position': int(self.settings_sheet['B4'].value),    # Начальная позиция
            'measure_unit': self.settings_sheet['B5'].value           # Единица измерения
        }

    def get_translations(self) -> Dict[str, str]:
        """
        Получает словарь переводов с листа 'Переводчик'
        
        Returns:
            Dict[str, str]: Словарь {английское_название: русское_название}
        """
        translations = {}
        
        for row in range(2, self.translator_sheet.max_row + 1):
            eng_name = self.translator_sheet[f'A{row}'].value
            rus_name = self.translator_sheet[f'B{row}'].value
            
            if eng_name and rus_name:
                translations[eng_name.strip()] = rus_name.strip()
        
        return translations

    def clear_work_sheet(self):
        """Очищает лист 'Рабочий'"""
        self.work_sheet.delete_rows(1, self.work_sheet.max_row)

    def write_product_data(self, products_data: List[Dict], settings: Dict, translations: Dict[str, str]):
        """
        Записывает данные о товарах в лист 'Рабочий' в формате CSV в столбец A
        
        Args:
            products_data (List[Dict]): Данные о товарах из PDF
            settings (Dict): Настройки из Excel
            translations (Dict[str, str]): Словарь переводов
        """
        current_position = settings['start_position']
        
        # Записываем заголовки в первую строку
        headers = ['ItemCode', 'Price', 'PositionId', 'Measure', 'Name']
        header_line = ','.join(headers)
        self.work_sheet.cell(row=1, column=1, value=header_line)
        
        # Записываем данные о товарах
        for row_num, product in enumerate(products_data, 2):
            if not product:
                continue
                
            # Вычисляем PositionId
            position_id = current_position
            
            # ItemCode: объединяем ARTICLE и FABRIC_CODE из PDF (без пробелов)
            item_code = f"{product['article']}{product['fabric_code']}".replace(' ', '')
            
            # ВЫЧИСЛЯЕМ ЦЕНУ ПО ПРАВИЛЬНОЙ ФОРМУЛЕ:
            # Цена в копейках = (PRICE_евро × Курс_евро × (1 + Наценка_процентов)) × 100
            price_eur = product['price_eur']
            exchange_rate = settings['exchange_rate']
            margin_percent = settings['margin_percent']
            
            # Применяем наценку: умножаем на (1 + margin_percent)
            price_with_margin = price_eur * exchange_rate * (1 + margin_percent)
            
            # Переводим в копейки (умножаем на 100) и берем целую часть (отбрасываем дробную)
            price_rub_kopecks = int(price_with_margin * 100)
            
            # Получаем название (перевод или оригинал)
            product_name = product['product_tipology']
            translated_name = translations.get(product_name, product_name)
            
            # Формируем строку CSV (без пробелов между элементами, кроме названия товара)
            csv_line = f"{item_code},{price_rub_kopecks},{position_id},{settings['measure_unit']},{translated_name}"
            
            # Записываем CSV строку в столбец A
            self.work_sheet.cell(row=row_num, column=1, value=csv_line)
            
            current_position += 1

    def save_workbook(self):
        """Сохраняет Excel файл"""
        try:
            self.wb.save(self.excel_path)
            return True
        except Exception as e:
            print(f"Ошибка при сохранении Excel файла: {e}")
            return False

    def set_cell_value(self, sheet_name: str, row: int, col: int, value):
        """
        Устанавливает значение ячейки в указанном листе
        
        Args:
            sheet_name (str): Название листа
            row (int): Номер строки
            col (int): Номер столбца
            value: Значение для записи
        """
        sheet = self.wb[sheet_name]
        sheet.cell(row=row, column=col, value=value)
